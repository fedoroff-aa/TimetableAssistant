import educational_objects
from docx import Document
from openpyxl import load_workbook, styles


class EducationalPrograms:
    # ATTENTION, PLEASE!
    # THERE ARE HIGH-RISK AREAS IN THE CODE BELOW!
    # SO DON'T LOOK AT THE CODE BELOW!

    # AUTHOR PROMISES THAT HE WILL WRITE A BETTER ONE
    # AS SOON AS HE GETS THE NORMAL PATTERN OF THE INPUT DATA
    # OR HE DECIDES TO WRITE A CUSTOMIZABLE VERSION.

    # TODO: FIX EVERYTHING К ЧЕРТЯМ СОБАЧЬИМ

    def __init__(self, path_to_file):

        self.path_to_file = path_to_file

    def get_programs(self):

        wb = load_workbook(self.path_to_file, data_only=True)
        classname_ranges_cols = [[(2, 4), (5, 7), (8, 16)], [(2, 10)], [(2, 11)]]
        classname_rows = [0, 2, 2]
        lesson_ranges_rows = [[(3, 24), (26, 41)], [(4, 25), (27, 39)], [(4, 26), (30, 43)]]
        letters_to_replace = [('C', 'С'), ('O', 'О'), ('P', 'Р'), ('E', 'Е'),
                              ('T', 'Т'), ('H', 'Н'), ('A', 'А'),
                              ('K', 'К'), ('X', 'Х'), ('B', 'В'), ('M', 'М')]
        programs = {}
        lessons = set()
        class_counter = {}

        for i in range(3):

            cols = wb[wb.sheetnames[i]].iter_cols()
            cols_c = 0
            lesson_names = {}

            for col in cols:

                current_classes = []
                current_lessons = {}
                rows_c = 0

                for current_cell in col[:50]:

                    cell_value = current_cell.value

                    if rows_c == classname_rows[i]:

                        if cell_value is not None:

                            if True in [x[0] <= cols_c < x[1] for x in classname_ranges_cols[i]]:

                                if ',' in cell_value:

                                    class_number = cell_value[:2] if cell_value[:2].isdigit() else cell_value[0]
                                    class_number = class_number.strip()

                                    class_letters = [x.strip() for x in
                                                     cell_value[len(class_number):].split(',')]

                                    current_classes = [class_number + x for x in class_letters]

                                else:

                                    current_classes = [cell_value.replace('(', ' ').split()[0].strip()]

                    elif True in [x[0] <= rows_c < x[1] for x in lesson_ranges_rows[i]] and cols_c == 1:

                        replaced_problem_letters_lesson = cell_value.strip().upper()

                        for letter_before, letter_after in letters_to_replace:
                            replaced_problem_letters_lesson = replaced_problem_letters_lesson \
                                .replace(letter_before, letter_after)

                        lesson_names[rows_c] = (cell_value.strip().upper() if
                                                all([not (1040 <= ord(x) <= 1103) for x in
                                                     cell_value.strip().upper()])
                                                else replaced_problem_letters_lesson)

                        lessons.add(lesson_names[rows_c])

                    elif (True in [x[0] <= rows_c < x[1] for x in lesson_ranges_rows[i]]
                          and cols_c >= 2 and str(cell_value).isdigit()):

                        together = current_cell.fill.start_color.index == 'FFFFFFFF'

                        if (lesson_names[rows_c], together) in current_lessons:

                            current_lessons[(lesson_names[rows_c], together)] =\
                                current_lessons[(lesson_names[rows_c], together)] + int(cell_value)

                        else:

                            current_lessons[(lesson_names[rows_c], together)] = int(cell_value)

                    rows_c += 1

                if current_lessons != {}:

                    for current_class in current_classes:

                        if current_class in class_counter:

                            class_counter[current_class] += 1
                            programs[current_class + str(class_counter[current_class])] = current_lessons

                        elif current_class not in class_counter:

                            class_counter[current_class] = 1
                            programs[current_class] = current_lessons
                cols_c += 1
        lessons = sorted(list(lessons))
        return programs, lessons


class Classrooms:

    def __init__(self, path_to_file):

        self.path_to_file = path_to_file

    def get_classrooms(self):

        data = Document(self.path_to_file).tables[0]
        classrooms = {}

        for i in range(1, len(data.rows)):

            classrooms[data.rows[i].cells[1].text] = educational_objects.Classroom(
                data.rows[i].cells[1].text,
                int(data.rows[i].cells[2].text)
            )

        return classrooms
